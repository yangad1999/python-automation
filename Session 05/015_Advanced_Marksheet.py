import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

os.system("cls")
marksheet_df = pd.read_csv("student_grades.csv")
pd.read
marksheet_df["Total"] = marksheet_df["Hindi"] + marksheet_df["Sanskrit"] + marksheet_df["English"] + marksheet_df["Social Science"] + marksheet_df["Maths"] + marksheet_df["Science"]

# Calculate average marks for each student
marksheet_df["Average"] = marksheet_df["Total"] / 6

# Define function to assign grades/results based on average marks
def assign_result(row):
    if any(row[subject] < 35 for subject in ["Hindi", "Sanskrit", "English", "Social Science", "Maths", "Science"]):
        return "Failed"
    elif row["Average"] >= 90:
        return "Merit"
    elif row["Average"] >= 60:
        return "First Class"
    elif row["Average"] >= 45:
        return "Second Class"
    elif row["Average"] >= 35:
        return "Third Class"
    else:
        return "Failed"

# Apply the function to create the 'Result' column
marksheet_df["Result"] = marksheet_df.apply(assign_result, axis=1)

# Save to Excel without formatting first
marksheet_df.to_excel("processed_marksheet.xlsx", index=False)

# Load the workbook and select the active worksheet
wb = load_workbook("processed_marksheet.xlsx")
ws = wb.active

# Define styles
header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
header_font = Font(bold=True)
center_alignment = Alignment(horizontal="center")
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Colors for results
colors = {
    "Failed": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Merit": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
    "First Class": PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid"),
    "Second Class": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
    "Third Class": PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
}

# Apply header style
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# Apply row styles and borders
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.alignment = center_alignment
        cell.border = thin_border
    result = row[-1].value
    if result in colors:
        fill = colors[result]
        for cell in row:
            cell.fill = fill

# Auto-fit column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the formatted workbook
wb.save("processed_marksheet.xlsx")

print(marksheet_df)
